import openpyxl as pxl
import math

exp = pxl.load_workbook('c:/Users/sagni/Downloads/avgcancer_sutirtha_c.xlsx')
sheet = exp['avgcancer_sutirtha_c']


def round_decimals_down(number: float, decimals: int = 2):
    """
    Returns a value rounded down to a specific number of decimal places.
    """
    if not isinstance(decimals, int):
        raise TypeError("decimal places must be an integer")
    elif decimals < 0:
        raise ValueError("decimal places has to be 0 or more")
    elif decimals == 0:
        return math.floor(number)

    factor = 10 ** decimals
    return math.floor(number * factor) / factor


def round_decimals_up(number: float, decimals: int = 2):
    """
    Returns a value rounded up to a specific number of decimal places.
    """
    if not isinstance(decimals, int):
        raise TypeError("decimal places must be an integer")
    elif decimals < 0:
        raise ValueError("decimal places has to be 0 or more")
    elif decimals == 0:
        return math.ceil(number)

    factor = 10 ** decimals
    return math.ceil(number * factor) / factor


peak = float(input("Enter the Peak Value: "))
lower_range = round_decimals_down(peak) - 0.1
upper_range = round_decimals_up(peak) + 0.1
result = []
mz_abs_arr = [[]]
flag = 0
print("------------------------------------------------------------------------------------------")
print('ANALYSIS FOR: Avg cancer samples')
print("------------------------------------------------------------------------------------------")
print('Format for the Result Array: [BC(i), Max_Intensity_Value, Correspoding_M/Z_Ratio_for_BC(i)]')
print("------------------------------------------------------------------------------------------")
for x in range(1, sheet.max_column+1, 2):
    # print("Finding for", sheet.cell(row=2, column=x+1).value)

    lowpos = None
    uppos = None
    result.append([sheet.cell(row=2, column=x+1).value])

    # print(lower_range,upper_range)
    for i in range(3, sheet.max_row+1):
        if float(sheet.cell(row=i, column=x).value) >= lower_range:
            lowpos = i
            break

    for i in range(lowpos+1, sheet.max_row+1):
        if float(sheet.cell(row=i, column=x).value) > upper_range:
            uppos = i-1
            break

    arr = []
    absint = []
    
    for i in range(lowpos, uppos+1):
        arr.append([float(sheet.cell(row=i, column=x).value),float(sheet.cell(row=i, column=x+1).value)])
        absint.append(float(sheet.cell(row=i, column=x+1).value))
        
    maxint = max(absint)

    absint.sort()
    for i in arr:
        if i[1] == maxint:

            result[-1].append(i[0])
            result[-1].append(i[1])
            mz_abs_arr[-1].append(i[0])
            mz_abs_arr[-1].append(i[1])
            #Comment these two lines and uncomment the Final three lines to print all the result_arrays at once 
            print("BC Cell Value: ",result[0][0])
            print("Maximum Intensity within Peak+-0.1:",result[0][1])
            print("M/Z Ratio to the corresponding Maximum Intensity:",result[0][2])
            flag = flag + 1
            
            result = []
            break

    arr.sort()
    sortarr = []

    for i in absint:
        for j in arr:
            if j[1] == i:
                sortarr.append([j[0], i])
                break
    print("------------------------------------------------------------------------------------------")
    print("Sorted array for:", sheet.cell(row=2, column=x+1).value)
        
    
    
    # Comment the below two Lines,  if you want to dont want to see the sorted array
    
    for z in sortarr:
        print(z)
        
    print("------------------------------------------------------------------------------------------")
    
print("Total Iterations Done: ",flag)

# print('Format for Priting will be: [ BC(i), Max_Intensity_Value_within_that_range, Correspoding_M/_Ratio_for_BC(i) ] ')
# for i in result:
#     print("For",i)